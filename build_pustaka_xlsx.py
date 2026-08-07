# -*- coding: utf-8 -*-
"""Bangun Daftar Studi Pustaka PRIMA+ OPSI 2026 dalam format Excel (.xlsx).

Output: outputs/opsi_prima_pustaka/Daftar_Pustaka_PRIMA_OPSI_2026.xlsx
Entri sama persis dengan daftar pustaka pada generator proposal & full paper,
ditambah kolom metadata hasil validasi citation-and-reference-validator (7/8/2026).
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join("outputs", "opsi_prima_pustaka")
OUT_PATH = os.path.join(OUT_DIR, "Daftar_Pustaka_PRIMA_OPSI_2026.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="EDEDED")
HEADER_FONT = Font(name="Times New Roman", bold=True, size=11)
BODY_FONT = Font(name="Times New Roman", size=11)
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

# (sitasi_apa, tahun, bidang, kategori, volume, halaman, doi_url, validasi, catatan)
ROWS = [
    ("Balai Bahasa Provinsi Kalimantan Tengah. (2026). Layanan Uji Kemahiran Berbahasa Indonesia (UKBI).",
     "2026", "Institusional", "Sumber resmi pemerintah",
     "-", "-", "https://balaibahasakalteng.kemendikdasmen.go.id/layanan/ukbi/",
     "OK", "Lembaga resmi Kemdikdasmen; pendukung pemutakhiran data kebahasaan."),
    ("Branch, R. M. (2009). Instructional Design: The ADDIE Approach. Springer.",
     "2009", "Teori / Metode", "Buku internasional",
     "-", "-", "https://doi.org/10.1007/978-0-387-09506-6",
     "Valid", "DOI hijau OpenALEX/Crossref; dasar model ADDIE. Rujukan inti."),
    ("Carter, R. (2003). Language awareness. ELT Journal, 57(1), 64-65.",
     "2003", "Kesadaran berbahasa", "Artikel jurnal",
     "57(1)", "64-65", "https://doi.org/10.1093/elt/57.1.64",
     "Valid", "Tahun benar = 2003 (mengakhiri ragam lama 2013/2023)."),
    ("Chaer, A., & Agustina, L. (2014). Sosiolinguistik: Perkenalan awal. Rineka Cipta.",
     "2014", "Sosiolinguistik", "Buku nasional",
     "-", "-", "-",
     "OK", "Buku acuan umum; tanpa DOI."),
    ("Crystal, D. (2011). Internet Linguistics: A Student Guide. Routledge.",
     "2011", "Linguistik internet", "Buku internasional",
     "-", "-", "-",
     "OK", "Buku acuan umum; tanpa DOI."),
    ("Fairclough, N. (1992). Critical Language Awareness. Longman.",
     "1992", "Kesadaran berbahasa", "Buku internasional",
     "-", "-", "-",
     "OK", "Penerbit benar Longman (bukan 'Logman')."),
    ("Garvin, P. L., & Mathiot, M. (1968). The urbanization of the Guarani language... In J. A. Fishman (Ed.), Readings in the Sociology of Language (pp. 365-374). De Gruyter Mouton.",
     "1968", "Loyalitas / sikap bahasa", "Bab buku",
     "-", "365-374", "https://doi.org/10.1515/9783110805376.365",
     "Valid", "DOI hijau di OpenALEX; penulis & halaman sesuai."),
    ("Kementerian Pendidikan Dasar dan Menengah. (2026). Informasi Satuan Pendidikan: MAN Kotawaringin Timur, NPSN 30201526.",
     "2026", "Konteks lokal", "Sumber resmi",
     "-", "-", "https://referensi.data.kemendikdasmen.go.id/snpmb/site/sekolah?npsn=30201526",
     "OK", "Data satuan pendidikan tempat penelitian."),
    ("Luo, Z. (2023). The effectiveness of gamified tools for foreign language learning (FLL): A systematic review. Behavioral Sciences, 13(4), 331.",
     "2023", "Gamifikasi bahasa", "Artikel jurnal (internasional)",
     "13(4)", "331", "https://doi.org/10.3390/bs13040331",
     "Valid", "DOI hijau; MDPI OA; sistematik review."),
    ("Parlindungan Siahaan, A., Aldy Pradana, M., Citra Chairani, D., Heriyani Erizal, A., & Margareta Lase, Y. (2024). Pengaruh era digital terhadap pemakaian bahasa Indonesia di kalangan remaja melalui media sosial. PENG: Jurnal Ekonomi dan Manajemen, 2(1), 879-885.",
     "2024", "Bahasa remaja digital", "Artikel jurnal (nasional)",
     "2(1)", "879-885", "https://teewanjournal.com/index.php/peng/article/view/1026",
     "Perlu catatan", "DOI aslinya mati → diganti URL artikel. Tahun 2024 benar (How-to-Cite). Situs jurnal beraset spam footer → rujukan pendukung."),
    ("Prasetyaningrum, R. (2024). Pengaruh media sosial terhadap gaya bahasa dalam penulisan bahasa Indonesia pada remaja. Jurnal Sosial Humaniora dan Pendidikan, 3(1), 127-134.",
     "2024", "Artikel SINTA", "Artikel jurnal (SINTA)",
     "3(1)", "127-134", "https://doi.org/10.55606/inovasi.v3i1.2734",
     "Valid", "DOI hijau di OpenALEX; konfirmasi vol/hal."),
    ("Shortt, M., Tilak, S., Kuznetcova, I., Martens, B., & Akinkuolie, B. (2021). Gamification in mobile-assisted language learning: A systematic review of Duolingo literature from public release of 2012 to early 2020. Computer Assisted Language Learning, 36(3), 517-554.",
     "2021", "Gamifikasi bahasa", "Artikel jurnal (internasional)",
     "36(3)", "517-554", "https://doi.org/10.1080/09588221.2021.1933540",
     "Valid", "DOI hijau; vol/issue/hal ditambahkan."),
    ("Sugiyono. (2019). Metode Penelitian dan Pengembangan: Research and Development. Alfabeta.",
     "2019", "Metode / R&D", "Buku pembelajaran",
     "-", "-", "-",
     "OK", "Buku metode R&D standar; tanpa DOI."),
    ("UKBI. (2026). Uji Kemahiran Berbahasa Indonesia. Kementerian Pendidikan Dasar dan Menengah.",
     "2026", "Institusional / kebahasaan", "Sumber resmi",
     "-", "-", "https://ukbi.kemendikdasmen.go.id/",
     "OK", "Portal UKBI Kemdikdasmen."),
    ("Widyaningrum, A. Y., Yumarnamto, M., & Prijambodo, V. L. (2020). Analisis resepsi remaja Kota Sampit mengenai keberagaman di media. WACANA: Jurnal Ilmiah Ilmu Komunikasi, 19(1), 51-61.",
     "2020", "Konteks lokal (Sampit)", "Artikel jurnal nasional",
     "19(1)", "51-61", "-",
     "OK", "Tanpa DOI; memperkuat konteks lokal."),
]

CAT_ORDER = ["Konsep berbahasa / Kesadaran berbahasa",
             "Loyalitas / sikap bahasa",
             "Sosiolinguistik",
             "Bahasa remaja",
             "Rajarasi mobile",
             "Konteks lokal",
             "Institusional / kebahasaan",
             "Metode / R&D",
             "Sumber resmi",]
CAT_ORDER = ["Konseptual", "Metode", "Konteks & Sumber", "Pendukung"]

def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()

    # Sheet 1: Daftar Pustaka (rinci + validasi)
    ws = wb.active
    ws.title = "Daftar Pustaka"
    ws.merge_cells("A1:J1")
    ws["A1"] = "DAFTAR PUSTAKA PRIMA+ — OPSI 2026 (Hasil Validasi Citation & Reference 2026-08-07)"
    ws["A1"].font = Font(name="Times New Roman", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["No", "Sitasi APA", "Tahun", "Cakupan Topik", "Kategori",
               "Volume", "Halaman", "DOI / URL", "Status Validasi", "Catatan"]
    hr = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for i, r in enumerate(ROWS, start=1):
        rr = hr + i
        vals = [i] + list(r)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 5, 6, 7, 9) else WRAP

    widths = [5, 60, 8, 22, 18, 9, 9, 40, 14, 50]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"

    # Sheet 2: Ringkasan statistik
    ws2 = wb.create_sheet("Ringkasan Validasi")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "RINGKASAN VALIDASI DAFTAR PUSTAKA PRIMA+"
    ws2["A1"].font = HEADER_FONT
    status = {"Valid / OK": 0, "Perlu catatan": 0}
    for r in ROWS:
        key = r[7]
        if key == "Valid":
            status["Valid / OK"] += 1
        elif key == "OK":
            status["Valid / OK"] += 1
        else:
            status["Perlu catatan"] += 1

    items = [
        ("Total entri", str(len(ROWS))),
        ("Metadata terverifikasi (DOI/URL)", "%d" % sum(1 for r in ROWS if r[7].startswith("Valid"))),
        ("Entri DOI/URL belum hijau", "%d" % sum(1 for r in ROWS if r[7] in ("Perlu catatan",))),
        ("Status keseluruhan", "Siap ke final; 1 entri (Parlindungan) tetap rujukan pendukung"),
    ]
    r0 = 3
    for k, (label, val) in enumerate(items):
        ws2.cell(row=r0 + k, column=1, value=label).font = HEADER_FONT
        ws2.cell(row=r0 + k, column=2, value=val).font = BODY_FONT

    wb.save(OUT_PATH)
    print("OK ->", OUT_PATH)

if __name__ == "__main__":
    main()